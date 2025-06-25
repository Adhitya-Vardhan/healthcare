# STEP 26: Implement /patients/upload
# File: app/api/patients.py

import uuid
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException , Request
from sqlalchemy.orm import Session
from app.core.deps import get_db, get_current_user
from app.schemas.patient import PatientUploadResponse
from app.models.models import Patient, FileUpload
from app.utils.encryption import encrypt_field
from openpyxl import load_workbook
from datetime import datetime
from app.schemas.patient import UploadStatus

from app.schemas.patient import PatientListResponse, PatientRow
from app.utils.encryption import decrypt_field
from sqlalchemy import func

from slowapi import Limiter
from app.core.rate_limitter import limiter

router = APIRouter()


@router.post("/patients/upload", response_model=PatientUploadResponse)
@limiter.limit("10/hour", key_func=lambda request: request.state.user_id)
def upload_patients(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Only managers can upload patients")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be Excel format")

    batch_id = f"batch_{uuid.uuid4().hex[:12]}"
    wb = load_workbook(file.file)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    required_headers = ["Patient ID", "First Name", "Last Name", "Date of Birth", "Gender"]
    if headers != required_headers:
        raise HTTPException(status_code=400, detail="Invalid column headers")

    total, success, failed = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        try:
            patient = Patient(
                patient_id=row[0],
                first_name_encrypted=encrypt_field(row[1]),
                last_name_encrypted=encrypt_field(row[2]),
                date_of_birth_encrypted=encrypt_field(str(row[3])),
                gender_encrypted=encrypt_field(row[4]),
                uploaded_by=current_user.id,
                encryption_key_version="v1.0",
                file_upload_batch_id=batch_id,
                data_hash="dummyhash"  # Optional: add hashing later
            )
            db.add(patient)
            success += 1
        except Exception:
            failed += 1
            continue
    db.commit()

    db.add(FileUpload(
        batch_id=batch_id,
        uploaded_by=current_user.id,
        original_filename=file.filename,
        file_size=file.spool_max_size,
        mime_type=file.content_type,
        total_records=total,
        successful_records=success,
        failed_records=failed,
        status="completed" if failed == 0 else "partial"
    ))
    db.commit()

    return PatientUploadResponse(
        batch_id=batch_id,
        filename=file.filename,
        total_records=total,
        status="completed" if failed == 0 else "partial",
        uploaded_at=datetime.utcnow()
    )

# STEP 29: Implement GET /patients/upload-status/{batchId}
# File: app/api/patients.py (extend)

@router.get("/upload-status/{batch_id}", response_model=UploadStatus)
def get_upload_status(batch_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Access denied")

    upload = db.query(FileUpload).filter(FileUpload.batch_id == batch_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Batch not found")

    return UploadStatus(
        batch_id=upload.batch_id,
        status=upload.status,
        total_records=upload.total_records,
        successful_records=upload.successful_records,
        failed_records=upload.failed_records,
        uploaded_at=upload.processing_started_at
    )

# STEP 31: Get paginated, decrypted patient data
# File: app/api/patients.py (extend)


@router.get("", response_model=PatientListResponse)
def list_patients(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    page: int = 1,
    limit: int = 20,
    search: str = None
):
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Access denied")

    query = db.query(Patient).filter(Patient.uploaded_by == current_user.id)

    total = query.count()
    pages = (total + limit - 1) // limit
    patients = query.offset((page - 1) * limit).limit(limit).all()

    result = []
    for p in patients:
        result.append(PatientRow(
            id=p.id,
            patient_id=p.patient_id,
            first_name=decrypt_field(p.first_name_encrypted),
            last_name=decrypt_field(p.last_name_encrypted),
            date_of_birth=decrypt_field(p.date_of_birth_encrypted),
            gender=decrypt_field(p.gender_encrypted),
            uploaded_by=current_user.username,
            uploaded_at=p.created_at,
            updated_at=p.updated_at
        ))

    return PatientListResponse(
        patients=result,
        page=page,
        limit=limit,
        total=total,
        pages=pages
    )


# STEP 33: Get single patient detail (decrypted)
# File: app/api/patients.py (extend)

from app.schemas.patient import PatientDetail

@router.get("/{patient_id}", response_model=PatientDetail)
def get_patient(patient_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Access denied")

    patient = db.query(Patient).filter(Patient.id == patient_id, Patient.uploaded_by == current_user.id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    return PatientDetail(
        id=patient.id,
        patient_id=patient.patient_id,
        first_name=decrypt_field(patient.first_name_encrypted),
        last_name=decrypt_field(patient.last_name_encrypted),
        date_of_birth=decrypt_field(patient.date_of_birth_encrypted),
        gender=decrypt_field(patient.gender_encrypted),
        uploaded_by=current_user.username,
        uploaded_at=patient.created_at,
        updated_at=patient.updated_at,
        batch_id=patient.file_upload_batch_id
    )


# STEP 34: Update patient record inline
# File: app/api/patients.py (extend)

from app.schemas.patient import UpdatePatientRequest

@router.put("/{patient_id}", response_model=PatientDetail)
def update_patient(patient_id: int, data: UpdatePatientRequest, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Access denied")

    patient = db.query(Patient).filter(Patient.id == patient_id, Patient.uploaded_by == current_user.id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    patient.first_name_encrypted = encrypt_field(data.first_name)
    patient.last_name_encrypted = encrypt_field(data.last_name)
    patient.date_of_birth_encrypted = encrypt_field(data.date_of_birth)
    patient.gender_encrypted = encrypt_field(data.gender)

    db.commit()
    db.refresh(patient)

    return PatientDetail(
        id=patient.id,
        patient_id=patient.patient_id,
        first_name=data.first_name,
        last_name=data.last_name,
        date_of_birth=data.date_of_birth,
        gender=data.gender,
        uploaded_by=current_user.username,
        uploaded_at=patient.created_at,
        updated_at=patient.updated_at,
        batch_id=patient.file_upload_batch_id
    )

# STEP 35: Delete a patient record
# File: app/api/patients.py (extend)

@router.delete("/{patient_id}")
def delete_patient(patient_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Access denied")

    patient = db.query(Patient).filter(Patient.id == patient_id, Patient.uploaded_by == current_user.id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    db.delete(patient)
    db.commit()
    return {"success": True, "message": "Patient record deleted successfully"}

# STEP 37: Search encrypted patients (exact match)
# File: app/api/patients.py (extend)

from app.schemas.patient import PatientSearchRequest

@router.post("/search", response_model=PatientListResponse)
def search_patients(
    criteria: PatientSearchRequest,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Access denied")

    query = db.query(Patient).filter(Patient.uploaded_by == current_user.id)

    if criteria.patient_id:
        query = query.filter(Patient.patient_id == criteria.patient_id)
    if criteria.first_name:
        query = query.filter(Patient.first_name_encrypted == encrypt_field(criteria.first_name))
    if criteria.last_name:
        query = query.filter(Patient.last_name_encrypted == encrypt_field(criteria.last_name))
    if criteria.date_of_birth:
        query = query.filter(Patient.date_of_birth_encrypted == encrypt_field(criteria.date_of_birth))
    if criteria.gender:
        query = query.filter(Patient.gender_encrypted == encrypt_field(criteria.gender))
    if criteria.date_range:
        query = query.filter(
            Patient.created_at.between(criteria.date_range["start"], criteria.date_range["end"])
        )

    matches = query.all()
    patients = [
        PatientRow(
            id=p.id,
            patient_id=p.patient_id,
            first_name=decrypt_field(p.first_name_encrypted),
            last_name=decrypt_field(p.last_name_encrypted),
            date_of_birth=decrypt_field(p.date_of_birth_encrypted),
            gender=decrypt_field(p.gender_encrypted),
            uploaded_by=current_user.username,
            uploaded_at=p.created_at,
            updated_at=p.updated_at,
        ) for p in matches
    ]

    return PatientListResponse(
        patients=patients,
        page=1,
        limit=len(patients),
        total=len(patients),
        pages=1
    )
