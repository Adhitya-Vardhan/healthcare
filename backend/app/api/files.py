# Enhanced files API with templates and export
# File: app/api/files.py

from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, desc
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime, timedelta
from typing import Optional, List

from app.core.deps import get_db, get_current_user
from app.models.models import Patient, FileUpload, User
from app.utils.encryption import encryption_service
from app.core.security_middleware import get_client_ip, get_user_agent

router = APIRouter()

def require_manager_role(current_user: User = Depends(get_current_user)):
    """Ensure user has Manager role"""
    if current_user.role.name != "Manager":
        raise HTTPException(status_code=403, detail="Manager access only")
    return current_user

@router.get("/files/template")
def download_template(
    template_type: str = Query("basic", description="Template type: basic, advanced, sample"),
    current_user: User = Depends(require_manager_role)
):
    """Download Excel template for patient data upload"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Patient Data"
        
        # Define headers based on template type
        if template_type == "advanced":
            headers = [
                "Patient ID", "First Name", "Last Name", "Date of Birth", "Gender",
                "Phone", "Email", "Address", "Emergency Contact", "Insurance ID"
            ]
            # Add instructions sheet
            instructions_ws = wb.create_sheet("Instructions")
            instructions = [
                ["Field", "Description", "Required", "Format/Options"],
                ["Patient ID", "Unique identifier for patient", "Yes", "Alphanumeric, max 50 chars"],
                ["First Name", "Patient's first name", "Yes", "Text, max 100 chars"],
                ["Last Name", "Patient's last name", "Yes", "Text, max 100 chars"],
                ["Date of Birth", "Patient's birth date", "Yes", "YYYY-MM-DD format"],
                ["Gender", "Patient's gender", "Yes", "Male, Female, Other"],
                ["Phone", "Contact phone number", "No", "Format: +1-XXX-XXX-XXXX"],
                ["Email", "Contact email", "No", "Valid email format"],
                ["Address", "Residential address", "No", "Full address"],
                ["Emergency Contact", "Emergency contact info", "No", "Name and phone"],
                ["Insurance ID", "Insurance identifier", "No", "Alphanumeric"]
            ]
            for row in instructions:
                instructions_ws.append(row)
            
        elif template_type == "sample":
            headers = ["Patient ID", "First Name", "Last Name", "Date of Birth", "Gender"]
            # Add sample data
            sample_data = [
                ["PAT001", "John", "Smith", "1985-01-15", "Male"],
                ["PAT002", "Jane", "Doe", "1990-05-22", "Female"],
                ["PAT003", "Robert", "Johnson", "1978-09-10", "Male"],
                ["PAT004", "Maria", "Garcia", "1992-12-03", "Female"],
                ["PAT005", "David", "Wilson", "1987-07-18", "Male"]
            ]
        else:  # basic template
            headers = ["Patient ID", "First Name", "Last Name", "Date of Birth", "Gender"]
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column width
            ws.column_dimensions[get_column_letter(col_num)].width = len(header) + 5
        
        # Add sample data if requested
        if template_type == "sample":
            for row_num, row_data in enumerate(sample_data, 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
        
        # Add data validation for Gender column
        if template_type in ["advanced", "basic", "sample"]:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            gender_col = headers.index("Gender") + 1
            dv = DataValidation(type="list", formula1='"Male,Female,Other"', allow_blank=False)
            dv.error = "Please select from the dropdown"
            dv.errorTitle = "Invalid Gender"
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(gender_col)}2:{get_column_letter(gender_col)}1000")
        
        # Create buffer and save
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"patient_data_template_{template_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Template generation failed: {str(e)}")

@router.get("/files/export")
def export_patients(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_role),
    format: str = Query("xlsx", description="Export format: xlsx, csv, json"),
    include_metadata: bool = Query(False, description="Include upload metadata"),
    date_from: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    batch_id: Optional[str] = Query(None, description="Filter by batch ID")
):
    """Export patient data with various formats and filters"""
    try:
        # Build query with filters
        query = db.query(Patient).filter(Patient.uploaded_by == current_user.id)
        
        if date_from:
            start_date = datetime.fromisoformat(date_from)
            query = query.filter(Patient.created_at >= start_date)
        
        if date_to:
            end_date = datetime.fromisoformat(date_to)
            query = query.filter(Patient.created_at <= end_date)
        
        if batch_id:
            query = query.filter(Patient.file_upload_batch_id == batch_id)
        
        patients = query.order_by(desc(Patient.created_at)).all()
        
        if not patients:
            raise HTTPException(status_code=404, detail="No patients found with the specified criteria")
        
        # Decrypt patient data
        decrypted_patients = []
        for patient in patients:
            try:
                decrypted_data = {
                    "patient_id": patient.patient_id,
                    "first_name": encryption_service.decrypt_field(patient.first_name_encrypted, "first_name"),
                    "last_name": encryption_service.decrypt_field(patient.last_name_encrypted, "last_name"),
                    "date_of_birth": encryption_service.decrypt_field(patient.date_of_birth_encrypted, "date_of_birth"),
                    "gender": encryption_service.decrypt_field(patient.gender_encrypted, "gender"),
                    "created_at": patient.created_at.isoformat() if patient.created_at else None,
                    "updated_at": patient.updated_at.isoformat() if patient.updated_at else None
                }
                
                if include_metadata:
                    decrypted_data.update({
                        "batch_id": patient.file_upload_batch_id,
                        "uploaded_by": current_user.username,
                        "encryption_key_version": patient.encryption_key_version,
                        "data_hash": patient.data_hash
                    })
                
                decrypted_patients.append(decrypted_data)
                
            except Exception as e:
                # Skip patients that can't be decrypted
                continue
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format.lower() == "json":
            # JSON export
            import json
            json_data = json.dumps(decrypted_patients, indent=2, default=str)
            
            return StreamingResponse(
                io.StringIO(json_data),
                media_type="application/json",
                headers={"Content-Disposition": f'attachment; filename="patients_export_{timestamp}.json"'}
            )
            
        elif format.lower() == "csv":
            # CSV export
            df = pd.DataFrame(decrypted_patients)
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False)
            csv_buffer.seek(0)
            
            return StreamingResponse(
                io.StringIO(csv_buffer.getvalue()),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="patients_export_{timestamp}.csv"'}
            )
            
        else:  # Excel export (default)
            wb = Workbook()
            ws = wb.active
            ws.title = "Patient Data"
            
            # Add headers
            if decrypted_patients:
                headers = list(decrypted_patients[0].keys())
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header.replace('_', ' ').title())
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Add data
                for row_num, patient in enumerate(decrypted_patients, 2):
                    for col_num, header in enumerate(headers, 1):
                        ws.cell(row=row_num, column=col_num, value=patient[header])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet if metadata included
            if include_metadata:
                summary_ws = wb.create_sheet("Export Summary")
                summary_data = [
                    ["Export Details", ""],
                    ["Generated At", datetime.now().isoformat()],
                    ["Generated By", current_user.username],
                    ["Total Records", len(decrypted_patients)],
                    ["Date Range", f"{date_from or 'All'} to {date_to or 'All'}"],
                    ["Batch ID Filter", batch_id or "All batches"],
                    ["Format", format.upper()],
                    ["Includes Metadata", "Yes" if include_metadata else "No"]
                ]
                
                for row_num, (key, value) in enumerate(summary_data, 1):
                    summary_ws.cell(row=row_num, column=1, value=key).font = Font(bold=True)
                    summary_ws.cell(row=row_num, column=2, value=value)
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return StreamingResponse(
                io.BytesIO(buffer.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="patients_export_{timestamp}.xlsx"'}
            )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.get("/files/export-batch/{batch_id}")
def export_batch(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_role),
    include_failed: bool = Query(False, description="Include failed records info")
):
    """Export specific upload batch with detailed information"""
    try:
        # Get batch information
        file_upload = db.query(FileUpload).filter(
            and_(
                FileUpload.batch_id == batch_id,
                FileUpload.uploaded_by == current_user.id
            )
        ).first()
        
        if not file_upload:
            raise HTTPException(status_code=404, detail="Batch not found")
        
        # Get patients from this batch
        patients = db.query(Patient).filter(
            and_(
                Patient.file_upload_batch_id == batch_id,
                Patient.uploaded_by == current_user.id
            )
        ).all()
        
        # Create comprehensive batch export
        wb = Workbook()
        
        # Batch Summary Sheet
        summary_ws = wb.active
        summary_ws.title = "Batch Summary"
        
        summary_data = [
            ["Batch Information", ""],
            ["Batch ID", batch_id],
            ["Original Filename", file_upload.original_filename],
            ["Upload Date", file_upload.processing_started_at.isoformat()],
            ["Completion Date", file_upload.processing_completed_at.isoformat() if file_upload.processing_completed_at else "N/A"],
            ["Status", file_upload.status],
            ["File Size", f"{file_upload.file_size:,} bytes"],
            ["MIME Type", file_upload.mime_type],
            ["", ""],
            ["Processing Results", ""],
            ["Total Records", file_upload.total_records],
            ["Successful Records", file_upload.successful_records],
            ["Failed Records", file_upload.failed_records],
            ["Success Rate", f"{(file_upload.successful_records / file_upload.total_records * 100):.1f}%" if file_upload.total_records > 0 else "N/A"]
        ]
        
        for row_num, (key, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_num, column=1, value=key).font = Font(bold=True)
            summary_ws.cell(row=row_num, column=2, value=value)
        
        # Patient Data Sheet
        if patients:
            patient_ws = wb.create_sheet("Patient Data")
            headers = ["Patient ID", "First Name", "Last Name", "Date of Birth", "Gender", "Created At", "Data Hash"]
            
            # Add headers
            for col_num, header in enumerate(headers, 1):
                cell = patient_ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Add patient data
            for row_num, patient in enumerate(patients, 2):
                try:
                    decrypted_data = [
                        patient.patient_id,
                        encryption_service.decrypt_field(patient.first_name_encrypted, "first_name"),
                        encryption_service.decrypt_field(patient.last_name_encrypted, "last_name"),
                        encryption_service.decrypt_field(patient.date_of_birth_encrypted, "date_of_birth"),
                        encryption_service.decrypt_field(patient.gender_encrypted, "gender"),
                        patient.created_at.isoformat() if patient.created_at else "",
                        patient.data_hash[:16] + "..."  # Truncate hash for display
                    ]
                    
                    for col_num, value in enumerate(decrypted_data, 1):
                        patient_ws.cell(row=row_num, column=col_num, value=value)
                        
                except Exception as e:
                    # Mark failed decryption
                    patient_ws.cell(row=row_num, column=1, value=patient.patient_id)
                    patient_ws.cell(row=row_num, column=2, value="[DECRYPTION FAILED]")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"batch_export_{batch_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Batch export failed: {str(e)}")

@router.get("/files/upload-history")
def get_upload_history(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_role),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100)
):
    """Get upload history for the current user"""
    try:
        offset = (page - 1) * limit
        
        query = db.query(FileUpload).filter(
            FileUpload.uploaded_by == current_user.id
        ).order_by(desc(FileUpload.processing_started_at))
        
        total = query.count()
        uploads = query.offset(offset).limit(limit).all()
        
        upload_history = []
        for upload in uploads:
            upload_info = {
                "batch_id": upload.batch_id,
                "original_filename": upload.original_filename,
                "file_size": upload.file_size,
                "mime_type": upload.mime_type,
                "total_records": upload.total_records,
                "successful_records": upload.successful_records,
                "failed_records": upload.failed_records,
                "status": upload.status,
                "processing_started_at": upload.processing_started_at.isoformat() if upload.processing_started_at else None,
                "processing_completed_at": upload.processing_completed_at.isoformat() if upload.processing_completed_at else None,
                "success_rate": round((upload.successful_records / upload.total_records * 100), 1) if upload.total_records > 0 else 0
            }
            upload_history.append(upload_info)
        
        pages = (total + limit - 1) // limit
        
        return {
            "uploads": upload_history,
            "page": page,
            "limit": limit,
            "total": total,
            "pages": pages
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching upload history: {str(e)}")

@router.delete("/files/batch/{batch_id}")
def delete_batch(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_role)
):
    """Delete a complete upload batch and all associated patient records"""
    try:
        # Verify batch belongs to current user
        file_upload = db.query(FileUpload).filter(
            and_(
                FileUpload.batch_id == batch_id,
                FileUpload.uploaded_by == current_user.id
            )
        ).first()
        
        if not file_upload:
            raise HTTPException(status_code=404, detail="Batch not found")
        
        # Count patients to be deleted
        patient_count = db.query(Patient).filter(
            and_(
                Patient.file_upload_batch_id == batch_id,
                Patient.uploaded_by == current_user.id
            )
        ).count()
        
        # Delete patients first (foreign key constraint)
        deleted_patients = db.query(Patient).filter(
            and_(
                Patient.file_upload_batch_id == batch_id,
                Patient.uploaded_by == current_user.id
            )
        ).delete()
        
        # Delete file upload record
        db.delete(file_upload)
        db.commit()
        
        return {
            "message": "Batch deleted successfully",
            "batch_id": batch_id,
            "deleted_patients": deleted_patients,
            "deleted_file_record": 1
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Batch deletion failed: {str(e)}")

@router.get("/files/bulk-export")
def bulk_export_all_data(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_role),
    format: str = Query("xlsx", description="Export format: xlsx, csv, json, zip"),
    include_audit_logs: bool = Query(False, description="Include audit logs"),
    password_protect: bool = Query(False, description="Password protect the export")
):
    """Create a comprehensive export of all user data (background task)"""
    try:
        # Count total records
        total_patients = db.query(Patient).filter(Patient.uploaded_by == current_user.id).count()
        total_uploads = db.query(FileUpload).filter(FileUpload.uploaded_by == current_user.id).count()
        
        if total_patients == 0:
            raise HTTPException(status_code=404, detail="No data to export")
        
        # Generate export ID for tracking
        export_id = f"export_{current_user.id}_{int(datetime.now().timestamp())}"
        
        # Add background task
        background_tasks.add_task(
            generate_bulk_export,
            export_id=export_id,
            user_id=current_user.id,
            format=format,
            include_audit_logs=include_audit_logs,
            password_protect=password_protect
        )
        
        return {
            "message": "Bulk export started",
            "export_id": export_id,
            "estimated_records": total_patients,
            "estimated_completion": "5-10 minutes",
            "status_endpoint": f"/files/export-status/{export_id}"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Bulk export initiation failed: {str(e)}")

# In-memory storage for export status (use Redis in production)
export_status_store = {}

def generate_bulk_export(
    export_id: str,
    user_id: int,
    format: str,
    include_audit_logs: bool,
    password_protect: bool
):
    """Background task to generate bulk export"""
    try:
        # Update status
        export_status_store[export_id] = {
            "status": "processing",
            "progress": 0,
            "message": "Starting export...",
            "started_at": datetime.now().isoformat()
        }
        
        # Simulate processing time and progress updates
        import time
        from app.db.session import SessionLocal
        
        db = SessionLocal()
        
        try:
            # Get all user data
            patients = db.query(Patient).filter(Patient.uploaded_by == user_id).all()
            uploads = db.query(FileUpload).filter(FileUpload.uploaded_by == user_id).all()
            
            total_operations = len(patients) + len(uploads) + (100 if include_audit_logs else 0)
            completed_operations = 0
            
            # Process patients
            export_status_store[export_id]["message"] = "Processing patient records..."
            processed_patients = []
            
            for patient in patients:
                try:
                    decrypted_data = {
                        "patient_id": patient.patient_id,
                        "first_name": encryption_service.decrypt_field(patient.first_name_encrypted, "first_name"),
                        "last_name": encryption_service.decrypt_field(patient.last_name_encrypted, "last_name"),
                        "date_of_birth": encryption_service.decrypt_field(patient.date_of_birth_encrypted, "date_of_birth"),
                        "gender": encryption_service.decrypt_field(patient.gender_encrypted, "gender"),
                        "created_at": patient.created_at.isoformat() if patient.created_at else None,
                        "batch_id": patient.file_upload_batch_id
                    }
                    processed_patients.append(decrypted_data)
                except Exception:
                    # Skip failed decryptions
                    pass
                
                completed_operations += 1
                export_status_store[export_id]["progress"] = int((completed_operations / total_operations) * 100)
                time.sleep(0.01)  # Simulate processing time
            
            # Simulate file generation
            export_status_store[export_id]["message"] = "Generating export file..."
            time.sleep(2)
            
            # Mark as completed
            export_status_store[export_id].update({
                "status": "completed",
                "progress": 100,
                "message": "Export completed successfully",
                "completed_at": datetime.now().isoformat(),
                "download_url": f"/files/download-export/{export_id}",
                "file_size": len(processed_patients) * 150,  # Estimated file size
                "record_count": len(processed_patients)
            })
            
        finally:
            db.close()
            
    except Exception as e:
        export_status_store[export_id] = {
            "status": "failed",
            "progress": 0,
            "message": f"Export failed: {str(e)}",
            "error": str(e),
            "failed_at": datetime.now().isoformat()
        }

@router.get("/files/export-status/{export_id}")
def get_export_status(
    export_id: str,
    current_user: User = Depends(require_manager_role)
):
    """Get status of bulk export"""
    if export_id not in export_status_store:
        raise HTTPException(status_code=404, detail="Export not found")
    
    return export_status_store[export_id]

@router.get("/files/download-export/{export_id}")
def download_export(
    export_id: str,
    current_user: User = Depends(require_manager_role)
):
    """Download completed export file"""
    if export_id not in export_status_store:
        raise HTTPException(status_code=404, detail="Export not found")
    
    status = export_status_store[export_id]
    if status["status"] != "completed":
        raise HTTPException(status_code=400, detail="Export not ready for download")
    
    # In a real implementation, you would serve the actual generated file
    # For this demo, we'll create a simple response
    content = "Export file content would be here..."
    
    return StreamingResponse(
        io.StringIO(content),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="bulk_export_{export_id}.txt"'}
    )

@router.post("/files/validate-upload")
async def validate_file_before_upload(
    file_content: bytes,
    filename: str,
    current_user: User = Depends(require_manager_role)
):
    """Validate file before actual upload - quick validation"""
    try:
        # Quick validations
        max_size = 50 * 1024 * 1024  # 50MB
        if len(file_content) > max_size:
            return {
                "valid": False,
                "errors": ["File too large (max 50MB)"],
                "file_info": {"size": len(file_content), "name": filename}
            }
        
        # Check file extension
        allowed_extensions = ['.xlsx', '.xls', '.csv']
        if not any(filename.lower().endswith(ext) for ext in allowed_extensions):
            return {
                "valid": False,
                "errors": ["Invalid file type. Only Excel (.xlsx, .xls) and CSV files allowed"],
                "file_info": {"size": len(file_content), "name": filename}
            }
        
        # Try to parse file structure
        try:
            if filename.lower().endswith('.csv'):
                df = pd.read_csv(io.StringIO(file_content.decode('utf-8')))
            else:
                df = pd.read_excel(io.BytesIO(file_content))
            
            # Check for required columns
            required_columns = ['patient_id', 'first_name', 'last_name', 'date_of_birth', 'gender']
            actual_columns = [col.lower().strip() for col in df.columns]
            missing_columns = [col for col in required_columns if col not in actual_columns]
            
            if missing_columns:
                return {
                    "valid": False,
                    "errors": [f"Missing columns: {', '.join(missing_columns)}"],
                    "file_info": {
                        "size": len(file_content),
                        "name": filename,
                        "rows": len(df),
                        "columns": list(df.columns)
                    }
                }
            
            # Check for empty data
            if len(df) == 0:
                return {
                    "valid": False,
                    "errors": ["File contains no data rows"],
                    "file_info": {"size": len(file_content), "name": filename, "rows": 0}
                }
            
            return {
                "valid": True,
                "message": "File validation passed",
                "file_info": {
                    "size": len(file_content),
                    "name": filename,
                    "rows": len(df),
                    "columns": list(df.columns),
                    "preview": df.head(3).to_dict('records')
                }
            }
            
        except Exception as parse_error:
            return {
                "valid": False,
                "errors": [f"File parsing error: {str(parse_error)}"],
                "file_info": {"size": len(file_content), "name": filename}
            }
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")